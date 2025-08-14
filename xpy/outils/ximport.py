import wx,os
import platform
import csv
import xlrd
import datetime
from xpy.outils import xformat
from openpyxl import load_workbook

def IsFile(nomFichier, mute=False):
    nomFichier = nomFichier.replace("/", "\\")
    if os.path.isfile(nomFichier):
        return True
    else:
        if not mute:
            wx.MessageBox("Fichier%s : non présent!"%(nomFichier))
        return False

def GetOneColCellsProp(ws,cell,typ=datetime.datetime):
    # retourne le nombre de cells non vide sous cell, la min et max
    nbCells = 0
    if cell:
        row = cell.row + 1  # start below the target cell
        col = cell.column
        values = []

        for r in range(row, ws.max_row + 1):
            wcell = ws.cell(row=r, column=col)
            if wcell.value:
                nbCells += 1
                # Try to parse as right type
                if isinstance(wcell.value, typ):
                    values.append(wcell.value)

        # Step 3: Compute min and max dates
        minVal = min(values) if values else None
        maxVal = max(values) if values else None
    else:
        mess = f"Found {nbCells} cells {str(typ)} below {cell.coordinate}"
        wx.MessageBox(mess,"Echec lecture")
        return None, None, None
    return nbCells, minVal, maxVal

def GetFirstCell(ws,text=None,nblig=10,nbcol=15):
    if text: text = text.lower()
    cell = None
    # Search within the defined range
    for row in ws.iter_rows(min_row=1,max_row=nblig,min_col=1,max_col=nbcol):
        for cell in row:
            if cell.value and (not text or text in str(cell.value).lower()):
                break
        else:
            continue
        break
    if not cell:
        mess = "Aucune cellule trouvée en haut d'écran contenant %s" % text
        mess += "\n recheche faite sur %d lignes et %d colonnes"%(nblig,nbcol)
        wx.MessageBox(mess, "ECHEC EXCEL")
    #print(text, f"Found in cell {cell.coordinate} with value: {cell.value}")
    return cell
        
def GetSheetNames(wk):
    # pour fichiers xlsx, lecture de la propriété
    return wk.sheetnames

def GetOneSheet(wk,sheetname):
    # récupére ws pour GetDonnées fichier.xlsx
    return wk[sheetname]

def GetNomsCols(ws,nbcol=10):
    cell = GetFirstCell(ws,'date')
    ridx = cell.row
    cidx = cell.column
    values = [ ws.cell(row=ridx,column=cidx + i).value for i in range(0, nbcol) ]
    return values

def GetNbRows(ws):
    return ws.max_row

def GetDonneesExcel(ws,minrow=1,maxrow=1000,mincol=1,maxcol=10):
    #get handle on existing file (xlsx only)
    lstDonnees = []
    # ajustement zone de cellules non vides, choix entête
    for values in ws.iter_rows(min_row=minrow,max_row=maxrow,min_col=mincol,max_col=maxcol,values_only=True,):
        sansNull = [x for x in values if x]
        # balaye jusqu'à trouver une ligne non vide
        if len(sansNull)>0:
            for cell in values:
                if cell:
                    break
                else:
                    # une colonne ignorée
                    mincol += 1
                    maxcol += 1
            break
        else:
            # une ligne ignorée
            minrow +=1

    #loop through range values
    for values in ws.iter_rows(min_row=minrow,max_row=maxrow,min_col=mincol,max_col=maxcol,values_only=True,):
        sansNull = [x for x in values if x]
        if len(sansNull)>0:
            lstDonnees.append(values)
    return lstDonnees

def OpenFile(nomFichier):
    # Teste l'ouverture fichier et retourne son type et son pointeur (si xlsx)
    if platform.system() == "Windows":
        nomFichier = nomFichier.replace("/", "\\")
    if not IsFile(nomFichier):
        return None, None
    typeFile, file = None, None
    lstNom = nomFichier.split('.')
    if lstNom[-1] in ('csv','txt'):
        typeFile = 'csv'
        try:
            file = open(nomFichier, "rt", encoding='utf8')
            file.close()
        except Exception as err:
            wx.MessageBox("Erreur d'accès au fichier\n\nAppel: %s\nErreur: %d, %s" % (
            nomFichier, err.args[0], err.args[1]))
        file = None
    elif lstNom[-1] == 'xlsx':
        typeFile = 'xlsx'
        file = load_workbook(filename=nomFichier)
    elif lstNom[-1] == 'xls':
        mess = "Ce fichier xls doit être transformé en xlsx"
        wx.MessageBox(mess,"IMPOSSIBLE")
    else:
        mess = "Le fichier n'est pas csv ou xlsx"
        wx.MessageBox(mess,"IMPOSSIBLE")
    return typeFile, file

def GetFichierXls(nomFichier,minrow=1,maxrow=1000,mincol=1,maxcol=10):
    if not IsFile(nomFichier):
        return []
    # pour anciennes versions de fichiers excel jusqu'à 2003
    """
    Si on veut accéder aux informations de formattage des cellules, il faut faire :
            myBook = xlrd.open_workbook('myFile.xls', formatting_info = True)
    myBook.sheets() : renvoie les feuilles du fichier sous forme de liste (objets xlrd.sheet.Sheet).
    myBook.sheet_names() : renvoie les noms des feuilles du fichier sous forme de liste.
    myBook.sheet_by_name('Feuill1') : renvoie la feuille de nom indiqué.

    Méthodes sur les objets Sheet (feuilles) :
    mySheet.name : le nom de la feuille.
    mySheet.nrows et mySheet.ncols : le nombre de lignes et de colonnes du fichier.
    mySheet.row(0) : renvoie la première ligne sous forme de liste des valeurs de type xlrd.sheet.Cell.
    mySheet.col(0) : renvoie la première colonne sous forme de liste des valeurs de type xlrd.sheet.Cell.
    mySheet.cell_type(5, 1) : le type de la valeur à la ligne 6 et la colonne 2 (origine à 0).

    mySheet.cell(5, 1) : la cellule à la ligne 6 et la colonne 2 (origine à 0), objet de type xlrd.sheet.Cell.

    Méthodes sur les objets xlrd.sheet.Cell :
    myCell.value : renvoie la valeur.
    myCell.xf_index : renvoie l'index de formattage, voir ci-dessous.
    """
    wk = xlrd.open_workbook(nomFichier)
    sheet_names = wk.sheet_names()
    ws = wk.sheet_by_index(0) # renvoie la première feuille.
    lstDonnees = []
    for row in range(minrow,maxrow):
        ligne = []
        sansNull = []
        for col in range(mincol,maxcol):
            try:
                value = ws.cell_value(row - 1, col - 1)
                cellType = ws.cell_type(row-1, col-1)
                if cellType == xlrd.sheet.XL_CELL_DATE:
                    value = datetime.datetime(*xlrd.xldate.xldate_as_tuple(value, wk.datemode))
            except: value = None
            ligne.append(value)
            if value: sansNull.append(value)
        if len(sansNull)>0:
            lstDonnees.append(ligne)
    return lstDonnees

def GetFichierXlsx(dicOptions):
    nomFichier = dicOptions.pop('nomFichier',None)
    maxcol = dicOptions.pop('maxcol',10)
    ixSheet = dicOptions.pop('ixSheet',0)

    # Ouverture du fichier
    try:
        (typeFichier, wk) = OpenFile(nomFichier)
        sheetNames = GetSheetNames(wk)
        sheetName = sheetNames[ixSheet]
        # activation de la feuille
        ws = GetOneSheet(wk,sheetName)
    except Exception as err:
        mess = "Le fichier %s n'a pas pu être ouvert correctement "% nomFichier
        mess += "\nerreur: %s"%err
        wx.MessageBox(mess, "ECHEC OUVERTURE")
        return

    lstCol = GetNomsCols(ws,maxcol)
    xformat.NormaliseNomChamps(lstCol)

    cellDate = GetFirstCell(ws,'date')
    if cellDate:
        minrow = cellDate.row + 1  # start below the target cell
        mincol = cellDate.column
        maxrow = ws.max_row
    else:
        mess = "Le fichier %s n'a pas cellule 'date'"% nomFichier
        wx.MessageBox(mess, "ECHEC OUVERTURE")
        return

    lstDonnees = []

    #loop through range values
    for values in ws.iter_rows(min_row=minrow,max_row=maxrow,min_col=mincol,
                               max_col=maxcol,values_only=True,):
        sansNull = [x for x in values if x]
        if len(sansNull)>0:
            lstDonnees.append(values)
    return lstDonnees

def GetFichierCsv(nomFichier,delimiter="\t",detect=True):
    if not IsFile(nomFichier):
        return []
    if platform.system() == "Windows":
        nomFichier = nomFichier.replace("/", "\\")
    # ouverture du fichier en lecture seule
    try:
        fichier = open(nomFichier, "rt",encoding='utf8')
    except Exception as err :
        wx.MessageBox("Erreur d'accès au fichier\n\nAppel: %s\nErreur: %d, %s"%(nomFichier,err.args[0],err.args[1]))
        return []
    # csv.reader est la fonction qui lit le fichier ouvert
    donnees = [x for x in csv.reader(fichier,delimiter=delimiter)]
    fichier.close()
    if detect and len(donnees[0]) == 1:
        # le séparateur n'etait pas le bon ou inexistant, essais avec ';'
        donnees = GetFichierCsv(nomFichier, delimiter=";",detect=False)
    return donnees

if __name__ == '__main__':
    app = wx.App(0)
    os.chdir("..")
    #donnees = GetFichierCsv("../Versions.txt")
    #donnees = GetFichierXls("c:/temp/FichierTest.xls")
    dicOptions = {'nomFichier':"C:\\Users\\jbrun\\Desktop\\bribes\CREDIT MUT RELEVE.xlsx",
                  'ixsheet':0}
    donnees = GetFichierXlsx(**dicOptions)
    print(donnees[0])
    app.MainLoop()

